"""
Spherical Harmonics Functions for GRACE and GRACE-FO Data

This module provides functions to compute and visualize Earth's gravity field 
based on GRACE and GRACE-FO satellite data using spherical harmonic synthesis. 
It includes support for generating gravity values across custom-defined 
coordinate grids, as well as rendering them into visual heatmaps.

Main Features:
--------------
- `spherical_harmonics_date`:
    Computes gravity acceleration and potential on a coordinate grid 
    for a single GRACE dataset (single date).
    
- `spherical_harmonics_baseline`:
    Averages gravity data across all dates to produce a baseline 
    gravity field for comparison.

- `render_single`:
    Renders a single gravity heatmap projected on a cropped world map 
    for the given spatial extent.

- `render_double`:
    Renders two gravity heatmaps side by side for comparing different 
    dates over the same spatial region.

Dependencies:
-------------
- `os`
- `tqdm`
- `openpyxl`
- `numpy`
- `pandas`
- `matplotlib`
- `scipy`
"""

# External Imports
import os
import tqdm
import openpyxl
import numpy as np
import pandas as pd
import cartopy.crs as ccrs
import matplotlib.pyplot as plt
import cartopy.feature as cfeature
from scipy.special import lpmv, factorial

# Define Spherical Harmonics Functions
def spherical_harmonics_date(data_array, lat_precis=30, lon_precis=60, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi), J2=False, save_xlsx=False, file_name='output.xlsx'):
    """
    Args:
        data_array: 3D array with shape (97, 97, 4) of C and S coefficients and their standard deviations
        lat_precis: amount of latitudinal coordinate points
        lon_precis: amount of longitudinal coordinate points
        lat_range: tuple of minimum and maximum lattitude in radians
        lon_range: tuple of minimum and maximum longitude in radians
        J2: include (True) or exclude (False)
        save_xlsx: create excel sheet (True) or not (False)
        file_name: data output name file (Always add .xlsx)
    
    Returns:
        earth_grid_pot: 2D array of gravitational potential [km^2/s^2] along a coordinate grid with x as longitude and y as latitude
        earth_grid_acc: 2D array of gravitational radial acceleration [km/s^2] along a coordinate grid with x as longitude and y as latitude
    
    Notes:
        always close the excel file before running
    """

    # Define Coordinate Grid
    colat_range = ((np.pi / 2) - lat_range[1], (np.pi / 2) - lat_range[0])
    colat_array = np.linspace(colat_range[0], colat_range[1], lat_precis)
    
    colon_range = (lon_range[0] + np.pi, lon_range[1] + np.pi)
    colon_array = np.linspace(colon_range[0], colon_range[1], lon_precis)

    # Define Constants
    mu = 3.9860044150e+5    # Earth's Gravitational Parameter in km^3/s^2
    Re = 6378.13630000      # Earth's Mean Radius in km
    r = Re
    
    # Degrees n (2 to 96) And Orders m (0 to 96)
    n_vals = np.arange(2, 97)[:, np.newaxis] # (95, 1)
    m_vals = np.arange(0, 97)[np.newaxis, :] # (1, 97)
    
    # Broadcast To (95, 97) Shape Automatically
    n = np.broadcast_to(n_vals, (95, 97))
    m = np.broadcast_to(m_vals, (95, 97))
    
    # Define Legendre Polynomials And Normalize
    valid_mask = m <= n
    delta_0_m = (m == 0).astype(np.float64)
    numer = (2 - delta_0_m) * ((2 * n) + 1) * factorial(n - m)
    denom = factorial(n + m)
    norm = np.zeros_like(n, dtype=np.float64)
    norm[valid_mask] = np.sqrt(numer[valid_mask] / denom[valid_mask])

    if not J2:
        norm[0, 0] = 0.0 # Remove J2 if not wanted

    # Pre-compute Legendre Matrix: Shape (lat_precis, 95, 97) And Apply Normalization
    weighted_lpmv = np.zeros((lat_precis, 95, 97))
    for j, colat in enumerate(tqdm.tqdm(colat_array, desc="Pre-Computing Normalized Legendre Polynomials Per Coordinate")):
        l_vals = lpmv(m, n, np.cos(colat))
        weighted_lpmv[j] = np.nan_to_num(l_vals, nan=0.0) * norm
    print()

    # Pre-Compute Trigonometric Relations (Shape = (lon_precis, 97))
    m_range = np.arange(97)
    cos_m_phi = np.cos(m_range[np.newaxis, :] * colon_array[:, np.newaxis])
    sin_m_phi = np.sin(m_range[np.newaxis, :] * colon_array[:, np.newaxis])

    # Format Data
    trimmed_data = data_array[2:, :, :-2] # Remove Degree 0 And 1 Rows And Standard Deviations
    C = trimmed_data[:, :, 0]
    S = trimmed_data[:, :, 1]
    
    # Initialise Data Lists
    earth_grid_pot = np.zeros((lat_precis, lon_precis))
    earth_grid_acc = np.zeros((lat_precis, lon_precis))

    # Run Sum Before To Save Memory
    n_plus_1 = n + 1

    # Run Over All Coordinates
    for j in tqdm.tqdm(range(lat_precis), desc="Processing Coordinates"):
        L_j = weighted_lpmv[j]
        
        # Weighted coefficients summed over degree n
        C_sum = np.sum(C * L_j, axis=0)      # Vector length 97
        S_sum = np.sum(S * L_j, axis=0)      # Vector length 97
        C_sum_a = np.sum(C * L_j * n_plus_1, axis=0)
        S_sum_a = np.sum(S * L_j * n_plus_1, axis=0)

        # Dot product for all longitudes at once
        pot_row = (cos_m_phi @ C_sum) + (sin_m_phi @ S_sum)
        acc_row = (cos_m_phi @ C_sum_a) + (sin_m_phi @ S_sum_a)

        earth_grid_pot[j, :] += (mu / Re) * pot_row + (mu / r)
        earth_grid_acc[j, :] += (mu / (Re * r)) * acc_row + (mu / (r**2))
    
    # Convert Acceleration Data From km/s2 To m/s2
    earth_grid_acc *= 1e3
    
    # Save Data To Excel File
    if save_xlsx:
        # Format Potential And Acceleration Data
        lat_array = np.linspace(lat_range[1], lat_range[0], lat_precis)
        lat_array_degree = np.degrees(lat_array)
        lat_array_degree = lat_array_degree.reshape(-1, 1)
        pot_first_column = np.vstack([["km2/s2"], lat_array_degree])
        acc_first_column = np.vstack([["m/s\u00B2"], lat_array_degree])

        pot_grid_data = np.vstack((np.degrees(colon_array), earth_grid_pot))
        pot_grid_data = np.hstack((pot_first_column, pot_grid_data))

        acc_grid_data = np.vstack((np.degrees(colon_array), earth_grid_acc))
        acc_grid_data = np.hstack((acc_first_column, acc_grid_data))

        # Define DataFrames For Excel Output
        df1 = pd.DataFrame(C)
        df2 = pd.DataFrame(S)
        df3 = pd.DataFrame(norm)
        df4 = pd.DataFrame(pot_grid_data)
        df5 = pd.DataFrame(acc_grid_data)

        # Create File Path
        file_path = os.path.join("Gravity_Maps/Output", file_name)

        # Save To Different Sheets In Excel Files
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df1.to_excel(writer, sheet_name="C-Coefficients", index=False)
            df2.to_excel(writer, sheet_name="S-Coefficients", index=False)
            df3.to_excel(writer, sheet_name="Normalization Factors", index=False)
            df4.to_excel(writer, sheet_name="Gravity Potential", index=False, header=False)
            df5.to_excel(writer, sheet_name="Gravity Acceleration", index=False, header=False)
        
        # Load the workbook to modify styles
        wb = openpyxl.load_workbook(file_path)

        # Apply formatting to the last two sheets
        for sheet_name in ["Gravity Potential", "Gravity Acceleration"]:
            ws = wb[sheet_name]

            # Bold the first row (headers)
            for cell in ws[1]:  # First row
                cell.font = openpyxl.styles.Font(bold=True)

            # Bold the first column (excluding header row)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = openpyxl.styles.Font(bold=True)

        # Save the modified file
        wb.save(file_path)

        # Print Completion Statement
        print(f"\n✅ Data Written To {file_name}")

    return earth_grid_acc

def spherical_harmonics_baseline(data_lib, sorted_date_lst, lat_precis=30, lon_precis=60, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi), J2=False, save_xlsx=False):
    '''
    Args:
        data_lib: list of dictionaries containing 'date' and 'data_array' keys for all dates
        sorted_date_lst: list of dates in the same order as data_lib to ensure correct matching
        lat_precis: amount of latitudinal coordinate points
        lon_precis: amount of longitudinal coordinate points
        lat_range: tuple of minimum and maximum lattitude in radians
        lon_range: tuple of minimum and maximum longitude in radians
        J2: include (True) or exclude (False)
        save_xlsx: create excel sheet (True) or not (False)

    Returns:
        average_acc: 2D array of average gravitational radial acceleration [m/s^2] across all dates along a coordinate grid with x as longitude and y as latitude
    
    Notes:
        always close the excel file before running
    '''

    # Define Coordinate Grid
    colat_range = ((np.pi / 2) - lat_range[1], (np.pi / 2) - lat_range[0])
    colat_array = np.linspace(colat_range[0], colat_range[1], lat_precis)
    
    colon_range = (lon_range[0] + np.pi, lon_range[1] + np.pi)
    colon_array = np.linspace(colon_range[0], colon_range[1], lon_precis)

    # Define Constants
    mu = 3.9860044150e+5 
    Re = 6378.13630000 
    r = Re
    
    # Degrees n (2 to 96) And Orders m (0 to 96)
    n_vals = np.arange(2, 97)[:, np.newaxis] # (95, 1)
    m_vals = np.arange(0, 97)[np.newaxis, :] # (1, 97)
    
    # Broadcast To (95, 97) Shape Automatically
    n = np.broadcast_to(n_vals, (95, 97))
    m = np.broadcast_to(m_vals, (95, 97))
    
    # Define Legendre Polynomials And Normalize
    valid_mask = m <= n
    delta_0_m = (m == 0).astype(np.float64)
    numer = (2 - delta_0_m) * ((2 * n) + 1) * factorial(n - m)
    denom = factorial(n + m)
    norm = np.zeros_like(n, dtype=np.float64)
    norm[valid_mask] = np.sqrt(numer[valid_mask] / denom[valid_mask])

    if not J2:
        norm[0, 0] = 0.0 # Remove J2 if not wanted

    # Pre-compute Legendre Matrix: Shape (lat_precis, 95, 97) And Apply Normalization
    weighted_lpmv = np.zeros((lat_precis, 95, 97))
    for j, colat in enumerate(tqdm.tqdm(colat_array, desc="Pre-Computing Normalized Legendre Polynomials Per Coordinate")):
        l_vals = lpmv(m, n, np.cos(colat))
        weighted_lpmv[j] = np.nan_to_num(l_vals, nan=0.0) * norm
    print()

    # Pre-Compute Trigonometric Relations (Shape = (lon_precis, 97))
    m_range = np.arange(97)
    cos_m_phi = np.cos(m_range[np.newaxis, :] * colon_array[:, np.newaxis])
    sin_m_phi = np.sin(m_range[np.newaxis, :] * colon_array[:, np.newaxis])

    # Extract Data
    data_dict = {entry['date']: entry['data_array'] for entry in data_lib}
    
    # Initialise Data Lists
    total_pot = np.zeros((lat_precis, lon_precis))
    total_acc = np.zeros((lat_precis, lon_precis))

    # Run Sum Before To Save Memory
    n_plus_1 = n + 1

    # Loop Over All Defined Dates
    for date in tqdm.tqdm(sorted_date_lst, desc="Processing Dates"):

        # Extract Data From Current Date
        data_array = data_dict[date]
        trimmed_data = data_array[2:, :, :-2]
        C, S = trimmed_data[:, :, 0], trimmed_data[:, :, 1]

        # Run Over All Coordinates
        for j in range(lat_precis):
            L_j = weighted_lpmv[j]
            
            # Weighted coefficients summed over degree n
            C_sum = np.sum(C * L_j, axis=0)      # Vector length 97
            S_sum = np.sum(S * L_j, axis=0)      # Vector length 97
            C_sum_a = np.sum(C * L_j * n_plus_1, axis=0)
            S_sum_a = np.sum(S * L_j * n_plus_1, axis=0)

            # Dot product for all longitudes at once
            pot_row = (cos_m_phi @ C_sum) + (sin_m_phi @ S_sum)
            acc_row = (cos_m_phi @ C_sum_a) + (sin_m_phi @ S_sum_a)

            total_pot[j, :] += (mu / Re) * pot_row + (mu / r)
            total_acc[j, :] += (mu / (Re * r)) * acc_row + (mu / (r**2))

    # Define Averages
    num_dates = len(sorted_date_lst)
    average_pot = total_pot / num_dates
    average_acc = (total_acc / num_dates) * 1e3 # Convert to m/s^2

    # Save Data To Excel File
    if save_xlsx:
        # Formatting for Excel
        lat_array_deg = np.degrees(np.linspace(lat_range[1], lat_range[0], lat_precis)).reshape(-1, 1)
        lon_array_deg = np.degrees(colon_array)

        def format_grid(data, unit):
            first_col = np.vstack([[unit], lat_array_deg.astype(str)])
            grid_body = np.vstack((lon_array_deg, data))
            return pd.DataFrame(np.hstack((first_col, grid_body)))

        df_pot = format_grid(average_pot, "km\u00B2/s\u00B2")
        df_acc = format_grid(average_acc, "m/s\u00B2")

        file_name = "Baseline Gravity Field.xlsx"
        file_path = os.path.join("Gravity_Maps/Output", file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_pot.to_excel(writer, sheet_name='Potential', index=False, header=False)
            df_acc.to_excel(writer, sheet_name='Acceleration', index=False, header=False)

        # Style formatting (OpenPyXL)
        wb = openpyxl.load_workbook(file_path)
        for name in ["Potential", "Acceleration"]:
            ws = wb[name]
            for cell in ws[1]: cell.font = openpyxl.styles.Font(bold=True)
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                for cell in row: cell.font = openpyxl.styles.Font(bold=True)
        wb.save(file_path)

        print(f"\n✅ Results saved to {file_name}")

    return average_acc

def render_single(earth_grid_acc, date, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi)):
    """
    Renders a gravity acceleration heatmap using Cartopy.
    
    Args:
        earth_grid_acc (np.ndarray): 2D array of values.
        date (str): Date string for the title/filename.
        lat_range/lon_range (tuple): Radians, defining the extent of earth_grid_acc.
    """
    # Convert Input Ranges From Radians To Degrees
    lon_deg = np.degrees(lon_range)
    lat_deg = np.degrees(lat_range)

    # Setup Plot And Projection
    projection = ccrs.PlateCarree()
    fig, ax = plt.subplots(figsize=(16, 8), subplot_kw={'projection': projection})

    # Set The Map Extent, i.e. Crop
    ax.set_extent([lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]], crs=ccrs.PlateCarree())

    # Add Background Features 
    ax.add_feature(cfeature.LAND, facecolor='lightgray')
    ax.add_feature(cfeature.OCEAN, facecolor='aliceblue')
    ax.add_feature(cfeature.COASTLINE, linewidth=0.8)
    ax.add_feature(cfeature.BORDERS, linestyle=':', alpha=0.5)

    # Overlay The Data (Extent: [min_lon, max_lon, min_lat, max_lat])
    im = ax.imshow(
        earth_grid_acc, 
        extent=[lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]],
        transform=ccrs.PlateCarree(),
        origin='upper', 
        cmap='coolwarm', 
        alpha=0.7,
        zorder=3 # Ensures data stays above the land/ocean colors
    )

    # Aesthetics
    plt.title(f"Gravity [{date}]", fontsize=15, pad=20)
    gl = ax.gridlines(draw_labels=True, dms=True, x_inline=False, y_inline=False, alpha=0.3)
    gl.top_labels = False
    gl.right_labels = False

    # Add Colorbar
    cbar = plt.colorbar(im, ax=ax, orientation='vertical', pad=0.02, aspect=30)
    cbar.set_label("Gravitational Acceleration [m/s\u00B2]", fontsize=12)

    # Save and Show
    filename = f"Gravity_Maps/Output/Gravity_Plot_{date}.png"
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Map successfully rendered and saved to {filename}")
    plt.show()

    return

def render_double(earth_grid_acc_1, earth_grid_acc_2, selected_dates, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi)):
    # Convert Radians to Degrees for mapping
    lon_deg = np.degrees(lon_range)
    lat_deg = np.degrees(lat_range)

    # Synchronize the Color Scale (vmin/vmax)
    vmin = min(earth_grid_acc_1.min(), earth_grid_acc_2.min())
    vmax = max(earth_grid_acc_1.max(), earth_grid_acc_2.max())

    # Setup Figure with PlateCarree projection
    projection = ccrs.PlateCarree()
    fig, axes = plt.subplots(1, 2, figsize=(16, 4), 
                             subplot_kw={'projection': projection})

    data_grids = [earth_grid_acc_1, earth_grid_acc_2]

    for i, ax in enumerate(axes):
        # Set The Extent (Crop)
        # This automatically handles the lat/lon window
        ax.set_extent([lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]], crs=ccrs.PlateCarree())
        
        # Add Geographic Layers
        ax.add_feature(cfeature.LAND, facecolor='lightgray', zorder=0)
        ax.add_feature(cfeature.OCEAN, facecolor='aliceblue', zorder=0)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.8, zorder=2)
        ax.add_feature(cfeature.BORDERS, linestyle=':', alpha=0.4, zorder=2)
        
        # Plot the Data
        im = ax.imshow(
            data_grids[i],
            extent=[lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]],
            transform=ccrs.PlateCarree(),
            origin='upper',
            alpha=0.7,
            cmap='coolwarm',
            vmin=vmin,
            vmax=vmax,
            zorder=1
        )
        
        # Title and Gridlines
        ax.set_title(f"Gravity [{selected_dates[i]}]", fontsize=14, pad=10)
        
        gl = ax.gridlines(draw_labels=True, dms=True, x_inline=False, y_inline=False, alpha=0.2)
        gl.top_labels = False
        gl.right_labels = False
        
        # Only show latitude labels on the left-most plot to clean up the center
        if i == 1:
            gl.left_labels = False

    # Shared Colorbar centered at the bottom
    cbar = fig.colorbar(im, ax=axes, orientation='horizontal', fraction=0.05, pad=0.08)
    cbar.set_label("Gravitational Acceleration [m/s\u00B2]", fontsize=12)

    # Save and Show
    filename = f"Gravity_Maps/Output/Gravity_Plot_{selected_dates[0]}_&_{selected_dates[1]}.png"
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Dual Heatmap Saved To {filename}")
    plt.show()

    return
