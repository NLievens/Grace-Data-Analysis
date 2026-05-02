
# External Imports
import os
import tqdm
import sys
import openpyxl
import datetime
import numpy as np
import pandas as pd
import cartopy.crs as ccrs
import matplotlib.pyplot as plt
import cartopy.feature as cfeature
from scipy.interpolate import interp1d
from scipy.special import lpmv, factorial

# Internal Imports
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from Data.Data_Reader import *
from Filters import *

# Define Delta Harmonic Coefficients
def compute_delta_harmonics(data_year_arr, date_year_arr, year_lst, max_order=96):
    '''
    Computes the delta spherical harmonic coefficients for C and S, applies necessary corrections (geocenter, SLR, GIA), and prepares the data for LSQR regression.

    Args:
        data_year_arr (np.ndarray): 5D array containing the original spherical harmonic coefficients and their standard deviations, indexed by year and month.
        date_year_arr (np.ndarray): 2D array containing the corresponding date strings for each year and month.
        year_lst (list): List of years to consider for the analysis.
        max_order (int): Maximum spherical harmonic degree/order to consider (default is 96).

    Returns:
        delta_t_lst (list): List of time deltas in days from the reference time t0.
        mask (np.ndarray): Boolean array indicating which coefficients are non-zero.
        org_tot_size (int): Original total number of coefficients before masking.
        CS_delta_vectors_arr (np.ndarray): Array of shape (num_months, num_coeffs) containing the delta coefficients for C and S.
        CS_std_delta_vectors_arr (np.ndarray): Array of shape (num_months, num_coeffs) containing the standard deviations of the delta coefficients for C and S.
    '''

    # Indexing & Pre-allocation
    year_index_arr = np.array(year_lst) - 2002
    considered_date_1D = date_year_arr[year_index_arr].flatten()    # Shape (288,) -> Years * 12 Months
    considered_data = data_year_arr[year_index_arr]                 # Shape (Years, 12, m_limit, m_limit, 4) -> (24, 12, 97, 97, 4) For max_order=96

    # Calculate t0
    first_valid = next((d for d in considered_date_1D if d.strip().lower() not in ("", "empty", "none")), None)
    if not first_valid:
        raise ValueError("❌ ERROR // No Valid Date Found")

    t0_year = datetime.strptime(first_valid.split("_")[0], "%d%m%Y").year   # 2002
    t0_time = datetime(t0_year, 1, 1)                                       # 2002-01-01 00:00:00

    # Initialize Lists
    data_C = []
    data_S = []
    data_C_std = []
    data_S_std = []
    delta_t_lst = []

    # Pre-Calculate Slice Limits
    m_limit = max_order + 1

    # Pre-Read GEO, SLR, and GIA Libraries
    geo_lib = extract_geocentric_coefficients()
    SLR_lib = extract_slr_coefficients()
    gia_C, gia_S = extract_GIA()

    print("✅ Read And Pre-Processed GEO, SLR, And GIA Data\n")

    # Process Each Month
    for i, month_date in enumerate(tqdm.tqdm(considered_date_1D, desc="Processing Months")):
        if month_date == "empty":
            continue

        # Calculate Delta t In Days
        current_dt = datetime.strptime(month_date.split("_")[0], "%d%m%Y")
        delta_t_lst.append((current_dt - t0_time).days)
        
        # Locate data (year index i // 12, month i % 12) (Only take the rows/cols up to max_order immediately)
        month_data = considered_data[i // 12, i % 12, :m_limit, :m_limit]
        
        # Split channels Of Original 97x97 Arrays
        C_orig, S_orig = month_data[:, :, 0].copy(), month_data[:, :, 1].copy()     # Shape (m_limit, m_limit) -> (97, 97) For max_order=96
        C_std, S_std   = month_data[:, :, 2].copy(), month_data[:, :, 3].copy()     # Shape (m_limit, m_limit) -> (97, 97) For max_order=96

        # Define Current Day Key
        current_key = month_date.split("_")[0][2:]

        # Add Geocenter Terms
        geo_vals = geo_lib.get(current_key)
        if geo_vals:
            C_orig[1, 0] = geo_vals['C10']
            C_orig[1, 1] = geo_vals['C11']
            S_orig[1, 1] = geo_vals['S11']

            C_std[1, 0] = geo_vals['C10_std']
            C_std[1, 1] = geo_vals['C11_std']
            S_std[1, 1] = geo_vals['S11_std']

        # Replace C20 And C30 With SLR Data
        SLR_vals = SLR_lib.get(current_key)
        if SLR_vals:
            C_orig[2, 0] = SLR_vals['C20']
            C_std[2, 0] = SLR_vals['C20_std']
            # Only replace C30 if it's not NaN (TN-14 has NaNs before 2012)
            if not np.isnan(SLR_vals['C30']):
                C_orig[3, 0] = SLR_vals['C30']
                C_std[3, 0] = SLR_vals['C30_std']
        
        # Remove Isostatic Rebound (GIA) (https://www.atmosp.physics.utoronto.ca/~peltier/data.php) (Inverse Sign Convention)
        dt_years = (current_dt - t0_time).days / 365.25

        C_orig += gia_C * dt_years
        S_orig += gia_S * dt_years

        # Filtering
        C_semi, S_semi = correlated_error_filter(C_orig, S_orig, max_order)
        C_fil, S_fil = gaussian_filter(C_semi, S_semi)

        # Append Data To Lists
        data_C.append(C_fil)
        data_S.append(S_fil)
        data_C_std.append(C_std)
        data_S_std.append(S_std)
    
    # Convert Lists To Arrays
    data_C_array = np.array(data_C)           # Shape (Months, 97, 97)
    data_S_array = np.array(data_S)           # Shape (Months, 97, 97)
    
    # Define Delta Values By Subtracting Mean
    C_mean = np.mean(data_C_array, axis=0)      # Shape (97, 97) -> Mean Across Months For Each Coefficient
    S_mean = np.mean(data_S_array, axis=0)      # Shape (97, 97) -> Mean Across Months For Each Coefficient

    delta_C = data_C_array - C_mean
    delta_S = data_S_array - S_mean

    # Create Global Mask
    flat_mean = np.concatenate([C_mean.ravel(), S_mean.ravel()])
    mask = (flat_mean != 0)
    org_tot_size = len(flat_mean)

    # Flatten & Concatenate
    CS_delta_vectors = []
    CS_std_delta_vectors = []
    for i in range(len(delta_C)):
        # Flatten C then S
        flat_CS = np.concatenate([delta_C[i].ravel(), delta_S[i].ravel()])
        flat_CS_std = np.concatenate([data_C_std[i].ravel(), data_S_std[i].ravel()])
        
        # Apply The Mask
        CS_delta_vectors.append(flat_CS[mask])
        CS_std_delta_vectors.append(flat_CS_std[mask])

    # Turn To Arrays
    CS_delta_vectors_arr = np.array(CS_delta_vectors)               # Shape (Months, Non-Zero Coefficients) -> (231, 9408) For max_order=96 (231, 97x97)
    CS_std_delta_vectors_arr = np.array(CS_std_delta_vectors)       # Shape (Months, Non-Zero Coefficients) -> (231, 9408) For max_order=96 (231, 97x97)

    return delta_t_lst, mask, org_tot_size, CS_delta_vectors_arr, CS_std_delta_vectors_arr

# Define Regression Model
def T_row(ti):
    '''
    Defines the regression model row for a given time delta ti. The model includes a constant term, linear and quadratic trends, and multiple periodic terms (annual, semi-annual, seasonal, 2-annual, and 5-annual).

    Args:
        ti (float): Time delta in days from the reference time t0.
        
    Returns:
        np.ndarray: A 1D array containing the regression model terms for the given time delta ti.
    '''

    return np.array([ 
        #1,                                              # 0 // Constant
        ti,                                             # 1 // Linear trend
        ti**2,                                          # 2 // Quadratic trend
        np.cos(2 * np.pi * ti / 365.2500),              # 3 // Annual (Cosine)
        np.sin(2 * np.pi * ti / 365.2500),              # 4 // Annual (Sine)
        np.cos(2 * np.pi * ti / 182.6250),              # 5 // Semi-Annual (Cosine)
        np.sin(2 * np.pi * ti / 182.6250),              # 6 // Semi-Annual (Sine)
        np.cos(2 * np.pi * ti / 91.3125),               # 7 // Seasonal (Cosine)
        np.sin(2 * np.pi * ti / 91.3125),               # 8 // Seasonal (Sine)
        np.cos(2 * np.pi * ti / (2 * 365.2500)),        # 9 // 2-Annual (Cosine)
        np.sin(2 * np.pi * ti / (2 * 365.2500)),        # 10 // 2-Annual (Sine)
        np.cos(2 * np.pi * ti / (5 * 365.2500)),        # 11 // 5-Annual (Cosine)
        np.sin(2 * np.pi * ti / (5 * 365.2500)),        # 12 // 5-Annual (Sine)
        np.cos(2 * np.pi * ti / (10 * 365.2500)),       # 13 // 10-Annual (Cosine)
        np.sin(2 * np.pi * ti / (10 * 365.2500))        # 14 // 10-Annual (Sine)
    ])

# Define Love Numbers
def get_love_numbers(max_degree=96):
    # Define Reference Values (https://doi.org/10.1029/98jb02844) (Table 1)
    l_values = np.array([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 30, 40, 50, 70, 100, 150, 200])
    k_l_values = np.array([0.000, 0.027, -0.303, -0.194, -0.132, -0.104, -0.089, -0.081,
                           -0.076, -0.072, -0.069, -0.064, -0.058, -0.051, -0.040, -0.033,
                           -0.027, -0.020, -0.014, -0.010, -0.007])
    
    # Set up interpolation function
    interp_func = interp1d(l_values, k_l_values, kind='linear', fill_value='extrapolate')
    
    # Generate interpolated Love numbers from l=0 to l_max
    l_full = np.arange(0, max_degree + 1)
    k_l_full = interp_func(l_full)
    
    return l_full, k_l_full

# Define Model Coefficients Through LSQR
def compute_model_coefficients(delta_t_lst, CS_delta_vectors, CS_std_delta_vectors, mask, org_tot_size, calc_uncertainty=False, max_order=96):
    '''
    Computes the model coefficients for each spherical harmonic coefficient using weighted least squares regression.

    Args:
        delta_t_lst (list): List of time deltas in days.
        CS_delta_vectors (np.ndarray): Array of shape (num_months, num_coeffs) containing the delta coefficients for C and S.
        CS_std_delta_vectors (np.ndarray): Array of shape (num_months, num_coeffs) containing the standard deviations of the delta coefficients for C and S.
        mask (np.ndarray): Boolean array indicating which coefficients are non-zero.
        org_tot_size (int): Original total number of coefficients before masking.
        calc_uncertainty (bool): Whether to calculate the covariance matrices for the model coefficients.
        max_order (int): Maximum spherical harmonic degree/order considered.

    Returns:
        coeff_models (np.ndarray): Array of shape (num_coeffs, num_params) containing the model coefficients for each spherical harmonic coefficient.
        SH_arr (np.ndarray): Array of shape (m_limit, m_limit, num_params, 2) containing the reconstructed C and S coefficients for each spherical harmonic degree/order.
        cov_SH_arr (np.ndarray or None): Array of shape (m_limit, m_limit, num_params, num_params, 2) containing the covariance matrices for the model coefficients of each spherical harmonic degree/order, or None if calc_uncertainty is False.
    '''

    # Setup Data
    delta_t_arr = np.array(delta_t_lst)
    Y = CS_delta_vectors.T                  # Shape (num_months, num_coeffs) -> Transpose to (num_coeffs, num_months)
    sigma = CS_std_delta_vectors.T          # Shape (num_months, num_coeffs) -> Transpose to (num_coeffs, num_months)
    weights = 1.0 / (sigma**2)              # Shape (num_coeffs, num_months) -> e.g. (9408, 231) For max_order=96
    
    num_coeffs, num_months = Y.shape
    T = np.vstack([T_row(ti) for ti in delta_t_arr])    # Shape (num_months, num_params) -> e.g. (231, 15)
    num_params = T.shape[1]
    TT = T.T                                            # Shape (num_params, num_months) -> e.g. (15, 231)

    coeff_models = np.zeros((num_coeffs, num_params))   # Shape (num_coeffs, num_params) -> e.g. (9408, 15) For max_order=96
    cov_beta_array = np.zeros((num_coeffs, num_params, num_params)) if calc_uncertainty else None

    # Weighted Least Squares Loop
    for i in range(num_coeffs):
        W_i = weights[i, :]
        XtWX = (TT * W_i) @ T
        try:
            XtWX_inv = np.linalg.inv(XtWX)
            coeff_models[i] = XtWX_inv @ (TT * W_i) @ Y[i]
            if calc_uncertainty:
                cov_beta_array[i] = XtWX_inv
        except np.linalg.LinAlgError:
            coeff_models[i] = np.linalg.lstsq(XtWX, (TT * W_i) @ Y[i], rcond=None)[0]

    # R-Squared
    Y_pred = (T @ coeff_models.T).T
    weighted_ssr = np.sum(weights * (Y - Y_pred)**2)
    Y_mean = np.mean(Y, axis=1, keepdims=True)
    weighted_sst = np.sum(weights * (Y - Y_mean)**2)
    r_squared = 1 - (weighted_ssr / weighted_sst)

    # Print Progress Statement
    print("✅ Model Coefficients Computed (R-Squared: {:.4f})".format(r_squared))

    # Reconstruct 4D SH_arr
    full_reconstructed = np.zeros((org_tot_size, num_params))
    full_reconstructed[mask] = coeff_models
    
    side = max_order + 1
    split = side * side
    C_rec = full_reconstructed[:split].reshape(side, side, num_params)
    S_rec = full_reconstructed[split:].reshape(side, side, num_params)
    SH_arr = np.stack([C_rec, S_rec], axis=-1)

    # Reconstruct 5D Covariance
    cov_SH_arr = None
    if calc_uncertainty:
        cov_SH_arr = np.zeros((side, side, num_params, num_params, 2))
        full_cov = np.zeros((org_tot_size, num_params, num_params))
        full_cov[mask] = cov_beta_array
        cov_SH_arr[:,:,:,:,0] = full_cov[:split].reshape(side, side, num_params, num_params)
        cov_SH_arr[:,:,:,:,1] = full_cov[split:].reshape(side, side, num_params, num_params)

    return coeff_models, SH_arr, cov_SH_arr

# Define EWH Grid Calculation
def compute_EWH_grid(SH_arr, elapsed_time, lat_precis=30, lon_precis=60, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi), J2=False, save_xlsx=False, file_name='output.xlsx'):
    '''
    Computes the Equivalent Water Height (EWH) grid from the spherical harmonic coefficients.
    
    Args:
        SH_arr (np.ndarray): Array of shape (m_limit, m_limit, num_params, 2) containing the reconstructed C and S coefficients for each spherical harmonic degree/order.
        elapsed_time (int): Time delta in days from the reference time t0 for which to compute the EWH grid.
        lat_precis (int): Latitude precision in points.
        lon_precis (int): Longitude precision in points.
        lat_range (tuple): Radians, defining the extent of the latitude range.
        lon_range (tuple): Radians, defining the extent of the longitude range.
        J2 (bool): Whether to include the J2 coefficient.
        save_xlsx (bool): Whether to save the output to an Excel file.
        file_name (str): Name of the Excel file to save (Always add .xlsx extension).

    Returns:
        earth_grid_pot (np.ndarray): 2D array of gravitational potential values on the defined grid.
        earth_grid_EWH (np.ndarray): 2D array of Equivalent Water Height values on the defined grid.
    '''

    # Define Constants
    mu = 3.9860044150e+5    # [km^3/s^2] Gravitational Parameter Of The Earth (WGS84) (Must Be In Same Units As Re For Potential Calculation)
    Re = 6378.13630000      # [km] Mean Radius Of The Earth (WGS84) (Must Be In Same Units As mu For Potential Calculation)
    r = Re

    rho_water = 1000        # [kg/m^3] Density Of Water
    rho_average = 5515      # [kg/m^3] Average Surface Density Of The Earth

    # Define Stokes Coefficients At Measurement Time
    T = T_row(elapsed_time)
    C_t = SH_arr[:, :, :, 0] @ T
    S_t = SH_arr[:, :, :, 1] @ T

    # Define Coordinate Grid
    colat_range = ((np.pi / 2) - lat_range[1], (np.pi / 2) - lat_range[0])
    colat_array = np.linspace(colat_range[0], colat_range[1], lat_precis)
    
    colon_range = (lon_range[0] + np.pi, lon_range[1] + np.pi)
    colon_array = np.linspace(colon_range[0], colon_range[1], lon_precis)

    # Degrees n (2 to 96) And Orders m (0 to 96)
    n_vals = np.arange(2, 97)[:, np.newaxis] # (95, 1)
    m_vals = np.arange(0, 97)[np.newaxis, :] # (1, 97)
    
    # Broadcast To (95, 97) Shape Automatically
    n = np.broadcast_to(n_vals, (95, 97))
    m = np.broadcast_to(m_vals, (95, 97))
    
    # Define Legendre Polynomials And Normalize
    valid_mask = m <= n
    delta_0_m = (m == 0).astype(np.float64)
    numer = (2 - delta_0_m) * ((2 * n) + 1) * factorial(n - m)
    denom = factorial(n + m)
    norm = np.zeros_like(n, dtype=np.float64)
    norm[valid_mask] = np.sqrt(numer[valid_mask] / denom[valid_mask])

    if not J2:
        norm[0, 0] = 0.0 # Remove J2 if not wanted

    # Pre-compute Legendre Matrix: Shape (lat_precis, 95, 97) And Apply Normalization
    weighted_lpmv = np.zeros((lat_precis, 95, 97))
    for j, colat in enumerate(tqdm.tqdm(colat_array, desc="Pre-Computing Normalized Legendre Polynomials Per Coordinate")):
        l_vals = lpmv(m, n, np.cos(colat))
        weighted_lpmv[j] = np.nan_to_num(l_vals, nan=0.0) * norm
    print()

    # Pre-Compute Trigonometric Relations (Shape = (lon_precis, 97))
    m_range = np.arange(97)
    cos_m_phi = np.cos(m_range[np.newaxis, :] * colon_array[:, np.newaxis])
    sin_m_phi = np.sin(m_range[np.newaxis, :] * colon_array[:, np.newaxis])

    # Format Data
    C_form = C_t[2:, :].reshape(95, 97)   # Shape (95, 97) -> Degrees 2-96, Orders 0-96 (Remove First 2 Degrees)
    S_form = S_t[2:, :].reshape(95, 97)   # Shape (95, 97) -> Degrees 2-96, Orders 0-96 (Remove First 2 Degrees)

    # Define Love Numbers
    degrees, love_numbers = get_love_numbers()

    # Apply Love Number Correction
    all_degrees = np.arange(97)
    love_corr_full = (2 * all_degrees + 1) / (1 + love_numbers)
    love_corr = love_corr_full[2:]      # Shape (95,) -> Remove First 2 Degrees To Match C_form And S_form Shapes

    C = C_form * love_corr[:, None]
    S = S_form * love_corr[:, None]

    # Initialise Data Lists
    earth_grid_pot = np.zeros((lat_precis, lon_precis))
    earth_grid_rho = np.zeros((lat_precis, lon_precis))

    # Run Over All Coordinates
    for j in tqdm.tqdm(range(lat_precis), desc="Processing Coordinates"):
        L_j = weighted_lpmv[j]
        
        # Weighted coefficients summed over degree n
        C_sum_pot = np.sum(C_form * L_j, axis=0)      # Vector length 97
        S_sum_pot = np.sum(S_form * L_j, axis=0)      # Vector length 97
        C_sum_rho = np.sum(C * L_j, axis=0)      # Vector length 97
        S_sum_rho = np.sum(S * L_j, axis=0)      # Vector length 97

        # Dot product for all longitudes at once
        pot_row = (cos_m_phi @ C_sum_pot) + (sin_m_phi @ S_sum_pot)
        mass_row = (cos_m_phi @ C_sum_rho) + (sin_m_phi @ S_sum_rho)

        earth_grid_pot[j, :] += (mu / Re) * pot_row + (mu / r)
        earth_grid_rho[j, :] += Re * (rho_average/3) * mass_row

    # Translate Surface Density Data [km*kg/m^3] To EWH [mm]
    earth_grid_EWH = (earth_grid_rho / rho_water) * 1e6

    # Save Data To Excel File
    if save_xlsx:
        # Format Potential And Equivalent Water Height Data
        lat_array = np.linspace(lat_range[1], lat_range[0], lat_precis)
        lat_array_degree = np.degrees(lat_array)
        lat_array_degree = lat_array_degree.reshape(-1, 1)
        pot_first_column = np.vstack([["km2/s2"], lat_array_degree])
        ewh_first_column = np.vstack([["mm"], lat_array_degree])

        pot_grid_data = np.vstack((np.degrees(colon_array), earth_grid_pot))
        pot_grid_data = np.hstack((pot_first_column, pot_grid_data))

        ewh_grid_data = np.vstack((np.degrees(colon_array), earth_grid_EWH))
        ewh_grid_data = np.hstack((ewh_first_column, ewh_grid_data))

        # Define DataFrames For Excel Output
        df1 = pd.DataFrame(C)
        df2 = pd.DataFrame(S)
        df3 = pd.DataFrame(norm)
        df4 = pd.DataFrame(pot_grid_data)
        df5 = pd.DataFrame(ewh_grid_data)

        # Create File Path
        file_path = os.path.join("LSQR_Analysis/Output", file_name)

        # Save To Different Sheets In Excel Files
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df1.to_excel(writer, sheet_name="C-Coefficients", index=False)
            df2.to_excel(writer, sheet_name="S-Coefficients", index=False)
            df3.to_excel(writer, sheet_name="Normalization Factors", index=False)
            df4.to_excel(writer, sheet_name="Gravity Potential", index=False, header=False)
            df5.to_excel(writer, sheet_name="Equivalent Water Height", index=False, header=False)
        
        # Load the workbook to modify styles
        wb = openpyxl.load_workbook(file_path)

        # Apply formatting to the last two sheets
        for sheet_name in ["Gravity Potential", "Equivalent Water Height"]:
            ws = wb[sheet_name]

            # Bold the first row (headers)
            for cell in ws[1]:  # First row
                cell.font = openpyxl.styles.Font(bold=True)

            # Bold the first column (excluding header row)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = openpyxl.styles.Font(bold=True)

        # Save the modified file
        wb.save(file_path)

        # Print Completion Statement
        print(f"\n✅ Data Written To {file_name}")

    return earth_grid_pot, earth_grid_EWH

# Define Single Plot Render
def render_single(earth_grid_EWH, date, sample_time, lat_range=(-np.pi/2, np.pi/2), lon_range=(-np.pi, np.pi)):
    """
    Renders a gravity acceleration heatmap using Cartopy.
    
    Args:
        earth_grid_acc (np.ndarray): 2D array of values.
        date (str): Date string for the title/filename.
        lat_range/lon_range (tuple): Radians, defining the extent of earth_grid_acc.
    """
    # Convert Input Ranges From Radians To Degrees
    lon_deg = np.degrees(lon_range)
    lat_deg = np.degrees(lat_range)

    # Setup Plot And Projection
    projection = ccrs.PlateCarree()
    fig, ax = plt.subplots(figsize=(16, 8), subplot_kw={'projection': projection})

    # Set The Map Extent, i.e. Crop
    ax.set_extent([lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]], crs=ccrs.PlateCarree())

    # Add Background Features 
    ax.add_feature(cfeature.LAND, facecolor='lightgray')
    ax.add_feature(cfeature.OCEAN, facecolor='aliceblue')
    ax.add_feature(cfeature.COASTLINE, linewidth=0.8)
    ax.add_feature(cfeature.BORDERS, linestyle=':', alpha=0.5)

    # Overlay The Data (Extent: [min_lon, max_lon, min_lat, max_lat])
    im = ax.imshow(
        earth_grid_EWH, 
        extent=[lon_deg[0], lon_deg[1], lat_deg[0], lat_deg[1]],
        transform=ccrs.PlateCarree(),
        origin='upper', 
        cmap='coolwarm', 
        alpha=0.7,
        zorder=3 # Ensures data stays above the land/ocean colors
    )

    # Aesthetics
    plt.title(f"LSQR Model At t = {sample_time} Days ({date[:2]}-{date[2:4]}-{date[4:]})")
    gl = ax.gridlines(draw_labels=True, dms=True, x_inline=False, y_inline=False, alpha=0.3)
    gl.top_labels = False
    gl.right_labels = False

    # Add Colorbar
    cbar = plt.colorbar(im, ax=ax, orientation='vertical', pad=0.02, aspect=30)
    cbar.set_label("Equivalent Water Height [mm]", fontsize=12)

    # Save and Show
    filename = f"LSQR_Analysis/Output/EWH_Plot_{date}.png"
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Map successfully rendered and saved to {filename}")
    plt.show()

    return
